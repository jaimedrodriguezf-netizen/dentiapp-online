import openpyxl

wb = openpyxl.load_workbook('./scratch/formulario033.xlsx', data_only=True)
sheet = wb.active

print("--- DETAILED COLUMNS FOR ROW 26 (E. Antecedentes Familiares) ---")
r = 26
c = 1
while c <= 64:
    cell = sheet.cell(row=r, column=c)
    coord = cell.coordinate
    
    merged_range = None
    for rng in sheet.merged_cells.ranges:
        if coord in rng:
            merged_range = rng
            break
            
    if merged_range:
        start_col = merged_range.min_col
        end_col = merged_range.max_col
        val = sheet.cell(row=r, column=start_col).value
        col_span = end_col - start_col + 1
        print(f"  Col {start_col}-{end_col} ({openpyxl.utils.get_column_letter(start_col)}-{openpyxl.utils.get_column_letter(end_col)}) [span {col_span}]: {val}")
        c = end_col + 1
    else:
        val = cell.value
        print(f"  Col {c} ({openpyxl.utils.get_column_letter(c)}) [no-merge]: {val}")
        c += 1
