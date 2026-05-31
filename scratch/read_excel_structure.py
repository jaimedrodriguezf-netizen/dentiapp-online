import openpyxl

wb = openpyxl.load_workbook('./scratch/formulario033.xlsx', data_only=True)
sheet = wb.active

print("--- EXCEL MERGED RANGES ---")
for r in sorted([str(rng) for rng in sheet.merged_cells.ranges]):
    print(r)

print("\n--- CELL VALUES BY ROW ---")
for r in range(1, 15):
    print(f"\nRow {r}:")
    row_info = []
    # Print only non-None or top-left merged cells to see actual layout
    for c in range(1, 25):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        coord = cell.coordinate
        
        # Check if it is top-left of a merged cell, or not merged at all
        is_merged = False
        is_top_left = False
        for rng in sheet.merged_cells.ranges:
            if coord in rng:
                is_merged = True
                if coord == rng.start_cell.coordinate:
                    is_top_left = True
                break
                
        if val is not None:
            status = " [MERGED START]" if is_top_left else (" [MERGED CHILD]" if is_merged else "")
            row_info.append(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {val}{status}")
    
    if row_info:
        for info in row_info:
            print("  " + info)
    else:
        print("  (Empty)")
