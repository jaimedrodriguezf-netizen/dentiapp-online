import openpyxl

wb = openpyxl.load_workbook('/home/jaimepop/dentiapp-online/scratch/formulario033.xlsx')
sheet = wb['1']

print("Tooth 48 detail (Rows 59-64, Cols G-I):")
for r in range(59, 65):
    for c in range(7, 10):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        border = cell.border
        border_str = []
        if border:
            if border.top and border.top.style: border_str.append(f"top:{border.top.style}")
            if border.bottom and border.bottom.style: border_str.append(f"bottom:{border.bottom.style}")
            if border.left and border.left.style: border_str.append(f"left:{border.left.style}")
            if border.right and border.right.style: border_str.append(f"right:{border.right.style}")
        border_desc = ", ".join(border_str) if border_str else "no-border"
        coord = openpyxl.utils.get_column_letter(c) + str(r)
        
        # Check if cell is merged
        is_merged = False
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                is_merged = True
                merged_coord = merged_range.coord
                break
        merged_desc = f"merged in {merged_coord}" if is_merged else "not-merged"
        
        print(f"Cell {coord}: val={val}, border={border_desc}, {merged_desc}")
