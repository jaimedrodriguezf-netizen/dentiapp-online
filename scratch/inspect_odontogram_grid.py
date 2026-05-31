import openpyxl

wb = openpyxl.load_workbook('/home/jaimepop/dentiapp-online/scratch/formulario033.xlsx')
sheet = wb['1']

print("Odontogram Grid Structure:")
for r in range(44, 65):
    row_cells = []
    # Loop from Col G (7) to Col BH (60)
    for c in range(7, 61):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        # If the cell is merged or has a value, show it
        border = cell.border
        border_str = ""
        if border:
            if border.top and border.top.style: border_str += "T"
            if border.bottom and border.bottom.style: border_str += "B"
            if border.left and border.left.style: border_str += "L"
            if border.right and border.right.style: border_str += "R"
        
        if val is not None or border_str:
            coord = openpyxl.utils.get_column_letter(c) + str(r)
            val_str = f"'{val}'" if val is not None else ""
            row_cells.append(f"{coord}:{val_str}({border_str})")
            
    if row_cells:
        print(f"Row {r:02d}: " + " | ".join(row_cells[:12]))
