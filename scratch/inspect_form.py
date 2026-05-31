import openpyxl

wb = openpyxl.load_workbook('/home/jaimepop/dentiapp-online/scratch/formulario033.xlsx')
print("Sheets:", wb.sheetnames)

# Inspect the active sheet or first sheet
sheet = wb['2']
print(f"\n--- Sheet 2 ---")
print(f"Dimensions: {sheet.dimensions}")

# Read first 120 rows and 20 columns
for r in range(1, 120):
    row_vals = []
    for c in range(1, 20):
        val = sheet.cell(row=r, column=c).value
        row_vals.append(str(val) if val is not None else "")
    # Print row if it has any content
    if any(row_vals):
        # Format the print to look readable
        print(f"Row {r:03d}: | " + " | ".join(row_vals[:15]))
