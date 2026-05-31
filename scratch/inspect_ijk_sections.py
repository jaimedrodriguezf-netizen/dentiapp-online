import openpyxl

wb = openpyxl.load_workbook('./scratch/formulario033.xlsx', data_only=True)
sheet = wb.active

# Sections I, J, K are in rows 65-86 based on form_structure.txt
# Full grid is A-BL = columns 1 to 64

print("=" * 120)
print("SECTIONS I, J, K — ROWS 65 to 86 — ALL 64 COLUMNS (A to BL)")
print("=" * 120)

# First: show merged ranges that touch rows 65-86
print("\n--- MERGED RANGES IN ROWS 65-86 ---")
for rng in sorted([str(r) for r in sheet.merged_cells.ranges]):
    # Parse row numbers from range string
    import re
    nums = re.findall(r'(\d+)', rng)
    if nums:
        row_nums = [int(n) for n in nums]
        # Check if any row in 65-86
        if any(65 <= r <= 86 for r in row_nums):
            print(f"  {rng}")

print("\n--- CELL VALUES ROW BY ROW ---")
for r in range(65, 87):
    print(f"\nRow {r}:")
    row_vals = []
    for c in range(1, 65):  # A=1 to BL=64
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        col_letter = openpyxl.utils.get_column_letter(c)
        coord = cell.coordinate

        is_merged = False
        is_top_left = False
        for rng in sheet.merged_cells.ranges:
            if coord in rng:
                is_merged = True
                if coord == rng.start_cell.coordinate:
                    is_top_left = True
                break

        if val is not None:
            tag = " [MERGE-START]" if is_top_left else (" [MERGED]" if is_merged else "")
            row_vals.append(f"  Col {c:2d} ({col_letter:>3s}): {repr(val)}{tag}")

    if row_vals:
        for v in row_vals:
            print(v)
    else:
        print("  (all empty)")

# Also check column widths for layout reference
print("\n--- COLUMN WIDTHS (A-BL) ---")
for c in range(1, 65):
    col_letter = openpyxl.utils.get_column_letter(c)
    dim = sheet.column_dimensions.get(col_letter)
    if dim and dim.width:
        print(f"  {col_letter}: width={dim.width}")
