import openpyxl

wb = openpyxl.load_workbook('/home/jaimepop/dentiapp-online/scratch/formulario033.xlsx')

def dump_sheet(sheet_name):
    sheet = wb[sheet_name]
    print(f"\n==========================================")
    print(f"SHEET: {sheet_name} (Dimensions: {sheet.dimensions})")
    print(f"==========================================")
    
    # Read row by row
    for r in range(1, 100):
        # Gather non-empty cells
        cells = []
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                cells.append(f"Col {openpyxl.utils.get_column_letter(c)}{r}: '{val}'")
        if cells:
            print(f"Row {r:02d}: " + " | ".join(cells))

dump_sheet('1')
dump_sheet('2')
